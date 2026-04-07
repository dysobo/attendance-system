from fastapi import FastAPI, Depends, HTTPException, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import PlainTextResponse, Response
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field
from typing import Optional, List, Literal, Union
from datetime import datetime, date, timedelta
from jose import JWTError, jwt
import database
import webhook as webhook_utils
import os
import hashlib
import requests
import json
import io
import re
import time
import xml.etree.ElementTree as ET
from hashlib import sha1
from collections import defaultdict
import calendar
from urllib.parse import quote

# ==================== 配置 ====================

SECRET_KEY = os.getenv("JWT_SECRET", "attendance-system-secret-key-2026")
ALGORITHM = "HS256"
TOKEN_EXPIRE_DAYS = 30

# 企业微信配置（从环境变量读取，用于回调验证）
WECHAT_CORP_ID = os.getenv("WECHAT_CORP_ID", "ww239ef8d083b060c3")
WECHAT_TOKEN = os.getenv("WECHAT_TOKEN", "b2bwx7eK4YTFcokbT8v")
WECHAT_ENCODING_AES_KEY = os.getenv("WECHAT_ENCODING_AES_KEY", "23aeqLKiM57mgOQrGLvON97UWNRQx6yYTJZGCaMeHCB")
WECHAT_AGENT_ID = int(os.getenv("WECHAT_AGENT_ID", "1000015"))
WECHAT_API_URL = os.getenv("WECHAT_API_URL", "http://ecs.dysobo.cn:56622")

def get_password_hash(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(plain_password, hashed_password):
    return get_password_hash(plain_password) == hashed_password

security = HTTPBearer()

app = FastAPI(title="考勤管理系统", version="5.0.0")

# CORS 配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== Pydantic 模型 ====================

class UserLogin(BaseModel):
    name: str
    password: str

class UserCreate(BaseModel):
    name: str
    password: str
    role: Literal["admin", "member"] = "member"
    phone: Optional[str] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[Literal["admin", "member"]] = None
    phone: Optional[str] = None
    wechat_user_id: Optional[str] = None
    enable_push: Optional[bool] = None

class ShiftCreate(BaseModel):
    user_id: int
    date: date
    shift_type: Literal["早班", "晚班", "休息"]
    note: Optional[str] = None

class ShiftUpdate(BaseModel):
    shift_type: Optional[Literal["早班", "晚班", "休息"]] = None
    note: Optional[str] = None

class TimeOffRequestCreate(BaseModel):
    date: date
    hours: float = Field(default=8.0, gt=0, multiple_of=0.5)
    type: Literal["U", "B", "S", "H", "C", "L", "J", "Y", "R", "N", "T", "Z"] = "U"
    reason: Optional[str] = None

class WebhookConfig(BaseModel):
    enabled: bool = False
    url: str = ""
    route_id: str = ""
    notify_time_off: bool = True
    notify_overtime: bool = True
    notify_time_off_approved: bool = False
    notify_overtime_approved: bool = False

class TimeOffRequestApprove(BaseModel):
    approved: bool
    admin_comment: Optional[str] = None

class OvertimeRecordCreate(BaseModel):
    date: date
    hours: float = Field(gt=0, multiple_of=0.5)
    reason: Optional[str] = None

class OvertimeRecordApprove(BaseModel):
    approved: bool
    admin_comment: Optional[str] = None

class MonthlyExportRequest(BaseModel):
    month: Optional[int] = None
    year: Optional[int] = None
    user_ids: List[int] = Field(default_factory=list)

# ==================== 工具函数 ====================

TIME_OFF_TYPE_NAMES = {
    "U": "调休",
    "B": "病假",
    "S": "事假",
    "H": "婚假",
    "C": "产假",
    "L": "护理假",
    "J": "经期假",
    "Y": "孕期假",
    "R": "哺乳假",
    "N": "年休假",
    "T": "探亲假",
    "Z": "丧假"
}

def create_access_token(data: dict):
    to_encode = data.copy()
    if 'sub' in to_encode:
        to_encode['sub'] = str(to_encode['sub'])
    to_encode.update({"exp": datetime.now().timestamp() + TOKEN_EXPIRE_DAYS * 86400})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def parse_date_value(value):
    if value in [None, "", "None"]:
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))

def parse_datetime_value(value):
    if value in [None, "", "None"]:
        return None
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value))

def build_textcard_title(category: str, status: str) -> str:
    return f"{category} | {status}"

def build_textcard_content(summary: str, lines: List[tuple[str, Optional[str]]], footer: Optional[str] = None) -> str:
    content_parts = [f'<div class="highlight">{summary}</div>']
    for label, value in lines:
        if value in [None, "", "None"]:
            continue
        content_parts.append(f'<div class="normal">{label}：{value}</div>')
    if footer:
        content_parts.append(f'<div class="gray">{footer}</div>')
    return "".join(content_parts)

def validate_half_hour_value(hours: float) -> float:
    try:
        normalized = float(hours)
    except (TypeError, ValueError):
        raise ValueError("时长格式不正确，请使用 0.5 小时为步长")

    if normalized <= 0 or round(normalized * 2) != normalized * 2:
        raise ValueError("时长必须大于 0，且以 0.5 小时为步长")
    return normalized

def parse_command_date(raw_value: str, today: Optional[date] = None) -> date:
    value = (raw_value or "").strip()
    reference_day = today or date.today()

    if value in ["今天", "今日"]:
        return reference_day
    if value == "昨天":
        return reference_day - timedelta(days=1)
    if value == "前天":
        return reference_day - timedelta(days=2)
    if value == "明天":
        return reference_day + timedelta(days=1)
    if value == "后天":
        return reference_day + timedelta(days=2)

    normalized = value.replace("/", "-").replace("年", "-").replace("月", "-").replace("日", "")
    for date_format in ("%Y-%m-%d", "%m-%d"):
        try:
            parsed = datetime.strptime(normalized, date_format).date()
            if date_format == "%m-%d":
                return parsed.replace(year=reference_day.year)
            return parsed
        except ValueError:
            continue

    raise ValueError("日期格式不支持，请使用 昨天 / 今天 / 明天 / 后天 / YYYY-MM-DD")

def parse_wechat_command(command_text: str, today: Optional[date] = None) -> dict:
    text = re.sub(r"\s+", " ", (command_text or "").strip())
    command_match = re.match(r"^(记加班|加班|调休)(?:申请)?\s*(.*)$", text)
    if not command_match:
        raise ValueError("暂不支持该指令")

    command_name = command_match.group(1)
    body = command_match.group(2).strip()
    if not body:
        raise ValueError("缺少时长信息，请使用“记加班 4h 今天 事由”或“调休 1h 今天 事由”")

    hours_match = re.search(r"(?P<hours>\d+(?:\.\d+)?)\s*(?:h|H|小时)", body)
    if not hours_match:
        raise ValueError("缺少时长信息，请使用 h 或 小时，例如 4h")

    hours = validate_half_hour_value(float(hours_match.group("hours")))
    body_without_hours = (body[:hours_match.start()] + " " + body[hours_match.end():]).strip()

    parsed_date = today or date.today()
    date_token = None
    for pattern in [
        r"(前天|昨天|今天|今日|明天|后天)",
        r"(\d{4}[/-]\d{1,2}[/-]\d{1,2})",
        r"(\d{1,2}[/-]\d{1,2})",
        r"(\d{4}年\d{1,2}月\d{1,2}日)",
        r"(\d{1,2}月\d{1,2}日)"
    ]:
        date_match = re.search(pattern, body_without_hours)
        if date_match:
            date_token = date_match.group(1)
            parsed_date = parse_command_date(date_token, today=today)
            body_without_hours = (body_without_hours[:date_match.start()] + " " + body_without_hours[date_match.end():]).strip()
            break

    reason = body_without_hours.strip(" ，,；;。")
    action = "overtime" if command_name in ["记加班", "加班"] else "time_off"

    return {
        "action": action,
        "date": parsed_date,
        "hours": hours,
        "reason": reason or None,
        "original_text": text
    }

def parse_wechat_approval_command(command_text: str) -> dict:
    text = re.sub(r"\s+", " ", (command_text or "").strip())
    approval_match = re.match(r"^(同意|不同意|拒绝)(?:\s+(.+))?$", text)
    if not approval_match:
        raise ValueError("暂不支持该审批指令")

    command_name = approval_match.group(1)
    admin_comment = (approval_match.group(2) or "").strip()
    return {
        "approved": command_name == "同意",
        "admin_comment": admin_comment or None,
        "original_text": text
    }

def is_supported_wechat_command(command_text: str) -> bool:
    text = re.sub(r"\s+", " ", (command_text or "").strip())
    return re.match(r"^(记加班|加班|调休)(?:申请)?(?:\s|$)", text) is not None or re.match(r"^(同意|不同意|拒绝)(?:\s|$)", text) is not None

def build_wechat_command_help_lines(user: Optional[database.User] = None) -> List[tuple[str, Optional[str]]]:
    lines = [
        ("加班", "记加班 4h 今天 设备调试"),
        ("调休", "调休 1h 明天 看牙")
    ]
    if user and user.role == "admin":
        lines.append(("审批", "同意 已处理；不同意 工时不足"))
    return lines

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(database.get_db)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user_id = int(user_id)
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    except (ValueError, TypeError):
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = db.query(database.User).filter(database.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

# ==================== 初始化 ====================

@app.on_event("startup")
def startup():
    database.init_db()
    db = database.SessionLocal()
    try:
        admin = db.query(database.User).filter(database.User.name == "admin").first()
        if not admin:
            admin = database.User(name="admin", password=get_password_hash("admin123"), role="admin")
            db.add(admin)
            db.commit()
            print("✅ 默认管理员已创建：admin / admin123")
    finally:
        db.close()

# ==================== API 路由 ====================

@app.get("/")
def root():
    return {"message": "考勤管理系统 API", "version": "5.0.0"}

# --- 用户认证 ---

@app.post("/api/login")
def login(user_data: UserLogin, db: Session = Depends(database.get_db)):
    user = db.query(database.User).filter(database.User.name == user_data.name).first()
    if not user or not verify_password(user_data.password, user.password):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    
    token = create_access_token({"sub": user.id, "role": user.role})
    return {
        "token": token,
        "user": {
            "id": user.id,
            "name": user.name,
            "role": user.role
        }
    }

@app.get("/api/me")
def get_me(current_user: database.User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "name": current_user.name,
        "role": current_user.role,
        "phone": current_user.phone,
        "wechat_user_id": current_user.wechat_user_id,
        "enable_push": current_user.enable_push
    }

@app.post("/api/users")
def create_user(user_data: UserCreate, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    existing = db.query(database.User).filter(database.User.name == user_data.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="用户名已存在")
    
    user = database.User(
        name=user_data.name,
        password=get_password_hash(user_data.password),
        role=user_data.role,
        phone=user_data.phone if hasattr(user_data, 'phone') else None
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"id": user.id, "name": user.name, "role": user.role, "phone": user.phone}

@app.get("/api/users")
def list_users(current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    users = db.query(database.User).all()
    return [{"id": u.id, "name": u.name, "role": u.role, "phone": u.phone, "wechat_user_id": u.wechat_user_id, "enable_push": u.enable_push} for u in users]

@app.post("/api/users/{user_id}/reset-password")
def reset_password(user_id: int, password_data: dict, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    user = db.query(database.User).filter(database.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user.password = get_password_hash(password_data.get("password", "123456"))
    db.commit()
    return {"message": "密码已重置"}

@app.post("/api/auth/change-password")
def change_password(password_data: dict, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    old_password = password_data.get("oldPassword")
    new_password = password_data.get("newPassword")
    
    if not old_password or not new_password:
        raise HTTPException(status_code=400, detail="请填写所有密码字段")
    
    if not verify_password(old_password, current_user.password):
        raise HTTPException(status_code=400, detail="当前密码错误")
    
    current_user.password = get_password_hash(new_password)
    db.commit()
    return {"message": "密码修改成功"}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="不能删除自己")
    
    user = db.query(database.User).filter(database.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    db.delete(user)
    db.commit()
    return {"message": "用户已删除"}

@app.put("/api/users/{user_id}/role")
def update_user_role(user_id: int, role_data: dict, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    user = db.query(database.User).filter(database.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    new_role = role_data.get("role")
    if new_role not in ["admin", "member"]:
        raise HTTPException(status_code=400, detail="无效的角色")
    
    user.role = new_role
    db.commit()
    db.refresh(user)
    return {"message": "角色已更新", "role": user.role}

@app.put("/api/users/{user_id}")
def update_user(user_id: int, user_data: UserUpdate, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    user = db.query(database.User).filter(database.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    if user_data.name is not None:
        existing = db.query(database.User).filter(database.User.name == user_data.name, database.User.id != user_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="用户名已存在")
        user.name = user_data.name
    if user_data.role is not None:
        user.role = user_data.role
    if user_data.phone is not None:
        user.phone = user_data.phone
    if user_data.wechat_user_id is not None:
        user.wechat_user_id = user_data.wechat_user_id
    if user_data.enable_push is not None:
        user.enable_push = user_data.enable_push
    
    db.commit()
    db.refresh(user)
    return {"message": "用户信息已更新", "user": {"id": user.id, "name": user.name, "role": user.role, "phone": user.phone, "wechat_user_id": user.wechat_user_id, "enable_push": user.enable_push}}

# ==================== 企业微信绑定 ====================

@app.get("/api/wechat/bind")
def get_wechat_bind(current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    return {
        "wechat_user_id": current_user.wechat_user_id,
        "enable_push": current_user.enable_push
    }

@app.post("/api/wechat/bind")
def bind_wechat(data: dict, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    current_user.wechat_user_id = data.get("wechat_user_id", "")
    current_user.enable_push = data.get("enable_push", True)
    db.commit()
    return {"message": "绑定成功", "wechat_user_id": current_user.wechat_user_id}

# ==================== 企业微信配置 ====================

@app.get("/api/wechat/config")
def get_wechat_config(current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    config = db.query(database.WechatConfig).first()
    if not config:
        return {
            "api_url": WECHAT_API_URL,
            "corp_id": WECHAT_CORP_ID,
            "secret": "",
            "agent_id": WECHAT_AGENT_ID,
            "token": WECHAT_TOKEN,
            "encoding_aes_key": WECHAT_ENCODING_AES_KEY,
            "enabled": False
        }
    
    return {
        "api_url": config.api_url,
        "corp_id": config.corp_id,
        "secret": config.secret,
        "agent_id": config.agent_id,
        "token": config.token,
        "encoding_aes_key": config.encoding_aes_key,
        "enabled": config.enabled
    }

@app.post("/api/wechat/config")
def save_wechat_config(config_data: dict, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    config = db.query(database.WechatConfig).first()
    if not config:
        config = database.WechatConfig()
        db.add(config)
    
    config.api_url = config_data.get("api_url", WECHAT_API_URL)
    config.corp_id = config_data.get("corp_id", WECHAT_CORP_ID)
    config.secret = config_data.get("secret", "")
    config.agent_id = config_data.get("agent_id", WECHAT_AGENT_ID)
    config.token = config_data.get("token", WECHAT_TOKEN)
    config.encoding_aes_key = config_data.get("encoding_aes_key", WECHAT_ENCODING_AES_KEY)
    config.enabled = config_data.get("enabled", False)
    config.updated_at = datetime.now()
    
    db.commit()
    return {"message": "配置已保存"}

@app.post("/api/wechat/test-push")
def test_wechat_push(data: dict, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    title = data.get("title", "测试推送")
    content = data.get("content", "这是一条测试消息")
    link_url = data.get("link", "https://x.dysobo.cn/kq/")
    
    success = send_wechat_message(current_user.id, title, content, link_url, db)
    
    if success:
        return {"message": "测试推送已发送"}
    else:
        raise HTTPException(status_code=500, detail="推送失败，请检查配置")

# ==================== 企业微信回调接口 ====================

import base64
import struct
from Crypto.Cipher import AES


def verify_wechat_signature(token: str, signature: str, timestamp: str, nonce: str, encrypt: str = None) -> bool:
    """验证企业微信签名

    签名计算：sha1(sort(token, timestamp, nonce, encrypt))
    GET 验证 URL 时 encrypt 是 echostr 参数
    POST 接收消息时 encrypt 是 XML 中的 Encrypt 字段
    """
    try:
        from hashlib import sha1
        sort_list = [token, timestamp, nonce]
        if encrypt:
            sort_list.append(encrypt)
        sort_list.sort()
        hash_str = ''.join(sort_list).encode('utf-8')
        hash_str = sha1(hash_str).hexdigest()
        return hash_str == signature
    except Exception as e:
        print(f"❌ 签名验证失败：{e}")
        return False


def decrypt_wechat_msg(encoding_aes_key: str, encrypted: str, corp_id: str) -> str:
    """解密企业微信消息（echostr 或消息体）

    官方格式：random(16B) + msg_len(4B) + msg + receiveid
    IV = aes_key[:16]
    """
    try:
        aes_key = base64.b64decode(encoding_aes_key + "=")
        ciphertext = base64.b64decode(encrypted)

        # IV 是 aes_key 的前 16 字节
        iv = aes_key[:16]
        cipher = AES.new(aes_key, AES.MODE_CBC, iv)
        decrypted = cipher.decrypt(ciphertext)

        # 去掉 PKCS#7 padding
        pad = decrypted[-1]
        if isinstance(pad, int):
            decrypted = decrypted[:-pad]
        else:
            decrypted = decrypted[:-ord(pad)]

        # 解析：random(16B) + msg_len(4B) + msg + receiveid
        msg_len = struct.unpack('>I', decrypted[16:20])[0]
        msg = decrypted[20:20 + msg_len]
        receiveid = decrypted[20 + msg_len:]

        if receiveid != corp_id.encode():
            print(f"⚠️ receiveid 不匹配：期望 {corp_id}，实际 {receiveid.decode('utf-8', errors='ignore')}")

        return msg.decode('utf-8')
    except Exception as e:
        print(f"解密失败：{e}")
        import traceback
        traceback.print_exc()
        return None


@app.get("/api/wechat/callback")
def wechat_callback_get(
    request: Request,
    msg_signature: str = "",
    timestamp: str = "",
    nonce: str = "",
    echostr: str = "",
    db: Session = Depends(database.get_db)
):
    """企业微信 URL 验证（GET 请求）
    
    验证流程：
    1. 企业微信发送 GET 请求，携带 msg_signature, timestamp, nonce, echostr
    2. echostr 是加密的，需要使用 EncodingAESKey 解密
    3. 签名计算：sha1(sort(token, timestamp, nonce, echostr))
    4. 验证签名通过后，解密 echostr 得到明文
    5. 返回解密后的明文（不能加引号，不能带 bom 头，不能带换行符）
    """
    # URL decode
    import urllib.parse
    query_string = request.scope.get("query_string", b"").decode()
    print(f"DEBUG: query_string={query_string}")
    for param in query_string.split("&"):
        if param.startswith("echostr="):
            echostr = urllib.parse.unquote(param[8:])
            print(f"DEBUG: echostr={echostr}")
            break

    # 获取配置
    config = db.query(database.WechatConfig).first()
    token = config.token if config and config.token else WECHAT_TOKEN
    encoding_aes_key = config.encoding_aes_key if config and config.encoding_aes_key else WECHAT_ENCODING_AES_KEY
    corp_id = config.corp_id if config and config.corp_id else WECHAT_CORP_ID
    
    print(f"🔍 回调验证请求:")
    print(f"  Token: [{token}]")
    print(f"  AES Key: [{encoding_aes_key[:10]}...]")
    print(f"  CorpID: [{corp_id}]")
    print(f"  请求签名：{msg_signature}")
    print(f"  Timestamp: {timestamp}")
    print(f"  Nonce: {nonce}")
    print(f"  Echostr: {echostr[:50]}...")
    
    # 验证签名（echostr 需要参与签名计算）
    if not verify_wechat_signature(token, msg_signature, timestamp, nonce, echostr):
        print(f"  ❌ 签名验证失败")
        raise HTTPException(status_code=403, detail="签名验证失败")
    
    print(f"  ✅ 签名验证成功")
    
    # 解密 echostr
    decrypted_msg = decrypt_wechat_msg(encoding_aes_key, echostr, corp_id)
    if decrypted_msg is None:
        print(f"  ❌ 解密失败")
        raise HTTPException(status_code=500, detail="解密失败")
    
    print(f"  ✅ 解密成功，返回明文：{decrypted_msg}")
    
    # 返回明文（不能加引号，不能带 bom 头，不能带换行符）
    return PlainTextResponse(content=decrypted_msg)


@app.post("/api/wechat/callback")
async def wechat_callback_post(
    request: Request,
    msg_signature: str = "",
    timestamp: str = "",
    nonce: str = "",
    db: Session = Depends(database.get_db)
):
    """接收企业微信消息（POST 请求）"""
    config = db.query(database.WechatConfig).first()
    token = config.token if config and config.token else WECHAT_TOKEN
    encoding_aes_key = config.encoding_aes_key if config and config.encoding_aes_key else WECHAT_ENCODING_AES_KEY
    corp_id = config.corp_id if config and config.corp_id else WECHAT_CORP_ID

    body = await request.body()
    body_str = body.decode('utf-8')
    print(f"📥 收到企业微信消息：{body_str[:200]}...")

    try:
        # 1. 从加密 XML 中提取 Encrypt 字段
        root = ET.fromstring(body_str)
        encrypt_node = root.find('Encrypt')
        if encrypt_node is None:
            print("❌ 未找到 Encrypt 字段")
            return 'success'

        encrypt_content = encrypt_node.text

        # 2. 验证签名：sha1(sort(token, timestamp, nonce, encrypt_content))
        if not verify_wechat_signature(token, msg_signature, timestamp, nonce, encrypt_content):
            print("❌ POST 消息签名验证失败")
            return 'success'

        # 3. 解密消息
        decrypted_xml = decrypt_wechat_msg(encoding_aes_key, encrypt_content, corp_id)
        if not decrypted_xml:
            print("❌ POST 消息解密失败")
            return 'success'

        print(f"📥 解密后消息：{decrypted_xml[:200]}...")

        # 4. 解析解密后的明文 XML
        msg_root = ET.fromstring(decrypted_xml)
        msg_type = msg_root.find('MsgType')
        event = msg_root.find('Event')
        event_key = msg_root.find('EventKey')
        from_user = msg_root.find('FromUserName')
        content_node = msg_root.find('Content')

        if msg_type is not None and msg_type.text == 'event':
            if event is not None and event.text == 'CLICK':
                if event_key is not None and event_key.text == 'PUSH_MY_STATS':
                    user_id = from_user.text if from_user is not None else ''
                    print(f'Click event: PUSH_MY_STATS from {user_id}')
                    try:
                        from wechat_click_handler import handle_push_stats
                        handle_push_stats(user_id, db)
                    except Exception as e:
                        print(f'Error calling handler: {e}')
        elif msg_type is not None and msg_type.text == 'text':
            wechat_user_id = from_user.text.strip() if from_user is not None and from_user.text else ""
            command_text = content_node.text.strip() if content_node is not None and content_node.text else ""
            bound_user = db.query(database.User).filter(database.User.wechat_user_id == wechat_user_id).first() if wechat_user_id else None
            if command_text and is_supported_wechat_command(command_text):
                print(f"📝 企业微信文本指令：{wechat_user_id} -> {command_text}")
                try:
                    handle_wechat_text_command(wechat_user_id, command_text, db)
                except ValueError as e:
                    print(f"⚠️ 企业微信指令处理失败：{e}")
                    send_wechat_command_feedback(
                        wechat_user_id=wechat_user_id,
                        success=False,
                        summary="指令未处理成功",
                        lines=[
                            ("原始内容", command_text),
                            ("原因", str(e))
                        ],
                        footer="请按示例格式发送",
                        db=db
                    )
                except Exception as e:
                    print(f"❌ 企业微信指令异常：{e}")
                    send_wechat_command_feedback(
                        wechat_user_id=wechat_user_id,
                        success=False,
                        summary="指令处理时发生异常",
                        lines=[
                            ("原始内容", command_text)
                        ],
                        footer="请稍后重试，或改用网页端提交",
                        db=db
                    )
            elif command_text:
                print(f"ℹ️ 未识别文本指令：{wechat_user_id} -> {command_text}")
                send_wechat_command_feedback(
                    wechat_user_id=wechat_user_id,
                    success=False,
                    summary="未能识别你的指令",
                    lines=[
                        ("原始内容", command_text),
                        *build_wechat_command_help_lines(bound_user)
                    ],
                    footer="请按以上格式发送",
                    db=db
                )
        return 'success'
    except Exception as e:
        print(f'Error parsing message: {e}')
        return 'success'


# ==================== 企业微信推送函数 ====================

def get_wechat_access_token(db: Session) -> str:
    config = db.query(database.WechatConfig).filter(database.WechatConfig.enabled == True).first()
    if not config or not config.corp_id or not config.secret:
        return ""
    
    try:
        url = f"{config.api_url}/cgi-bin/gettoken"
        params = {
            "corpid": config.corp_id,
            "corpsecret": config.secret
        }
        response = requests.get(url, params=params, timeout=10)
        result = response.json()
        
        if result.get("errcode") == 0:
            return result.get("access_token", "")
        else:
            print(f"❌ 获取 access_token 失败：{result}")
            return ""
    except Exception as e:
        print(f"❌ 获取 access_token 异常：{e}")
        return ""

def send_wechat_message_to_wechat_user(wechat_user_id: str, title: str, content: str, link_url: str = "", db: Session = None):
    if not db or not wechat_user_id:
        return False

    config = db.query(database.WechatConfig).filter(database.WechatConfig.enabled == True).first()
    if not config:
        return False

    access_token = get_wechat_access_token(db)
    if not access_token:
        return False

    try:
        url = f"{config.api_url}/cgi-bin/message/send?access_token={access_token}"
        payload = {
            "touser": wechat_user_id,
            "msgtype": "textcard",
            "agentid": config.agent_id,
            "textcard": {
                "title": title,
                "description": content,
                "url": link_url if link_url else "https://x.dysobo.cn/kq/",
                "btntxt": "查看详情"
            }
        }

        response = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=10)
        result = response.json()

        if result.get("errcode") == 0:
            print(f"✅ 企业微信消息发送成功：{wechat_user_id} - {title}")
            return True
        else:
            print(f"❌ 企业微信消息发送失败：{result}")
            return False
    except Exception as e:
        print(f"❌ 企业微信消息发送异常：{e}")
        return False

def send_wechat_message(user_id: int, title: str, content: str, link_url: str = "", db: Session = None):
    if not db:
        return False

    user = db.query(database.User).filter(database.User.id == user_id).first()
    if not user or not user.enable_push or not user.wechat_user_id:
        return False

    return send_wechat_message_to_wechat_user(user.wechat_user_id, title, content, link_url, db)

def create_time_off_request_for_user(
    user: database.User,
    request_date: date,
    hours: float,
    reason: Optional[str],
    db: Session,
    request_type: str = "U"
):
    request = database.TimeOffRequest(
        user_id=user.id,
        date=request_date,
        hours=hours,
        type=request_type,
        reason=reason
    )
    db.add(request)
    db.commit()
    db.refresh(request)

    type_name = TIME_OFF_TYPE_NAMES.get(request_type, "调休")
    link_url = "https://x.dysobo.cn/kq/?page=timeoff"
    title = build_textcard_title(f"{type_name}申请", "待审批")
    content = build_textcard_content(
        summary=f"{user.name} 提交了{type_name}申请",
        lines=[
            ("日期", str(request.date)),
            ("时长", f"{request.hours} 小时"),
            ("事由", request.reason)
        ],
        footer="请及时审批处理"
    )
    admins = db.query(database.User).filter(database.User.role == "admin").all()
    for admin in admins:
        send_wechat_message(admin.id, title, content, link_url, db)

    webhook_config = webhook_utils.load_webhook_config()
    if webhook_config.get("enabled") and webhook_config.get("notify_time_off"):
        webhook_utils.notify_time_off_request(
            user_name=user.name,
            date=str(request.date),
            hours=request.hours,
            reason=request.reason or "",
            webhook_config=webhook_config,
            link_url=f"https://x.dysobo.cn/kq/?page=timeoff&id={request.id}"
        )

    return request

def create_overtime_record_for_user(
    user: database.User,
    record_date: date,
    hours: float,
    reason: Optional[str],
    db: Session
):
    record = database.OvertimeRecord(
        user_id=user.id,
        date=record_date,
        hours=hours,
        reason=reason
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    link_url = "https://x.dysobo.cn/kq/?page=overtime"
    title = build_textcard_title("加班记录", "待确认")
    content = build_textcard_content(
        summary=f"{user.name} 提交了加班申请",
        lines=[
            ("日期", str(record.date)),
            ("时长", f"{record.hours} 小时"),
            ("事由", record.reason)
        ],
        footer="请及时确认处理"
    )
    admins = db.query(database.User).filter(database.User.role == "admin").all()
    for admin in admins:
        send_wechat_message(admin.id, title, content, link_url, db)

    webhook_config = webhook_utils.load_webhook_config()
    if webhook_config.get("enabled") and webhook_config.get("notify_overtime"):
        webhook_utils.notify_overtime_request(
            user_name=user.name,
            date=str(record.date),
            hours=record.hours,
            reason=record.reason or "",
            webhook_config=webhook_config,
            link_url=f"https://x.dysobo.cn/kq/?page=overtime&id={record.id}"
        )

    return record

def approve_time_off_request_record(
    request: database.TimeOffRequest,
    approver: database.User,
    approved: bool,
    admin_comment: Optional[str],
    db: Session
):
    request.status = "approved" if approved else "rejected"
    request.approved_by = approver.id
    request.admin_comment = admin_comment
    request.updated_at = datetime.now()
    db.commit()

    type_name = TIME_OFF_TYPE_NAMES.get(request.type, "调休")
    status_text = "已批准" if approved else "已拒绝"
    link_url = "https://x.dysobo.cn/kq/?page=timeoff"
    applicant = db.query(database.User).filter(database.User.id == request.user_id).first()
    applicant_name = applicant.name if applicant else "未知"

    title = build_textcard_title(f"{type_name}申请", status_text)
    content = build_textcard_content(
        summary=f"您的{type_name}申请{status_text}",
        lines=[
            ("日期", str(request.date)),
            ("时长", f"{request.hours} 小时"),
            ("审批意见", admin_comment)
        ],
        footer="点击查看详情"
    )
    send_wechat_message(request.user_id, title, content, link_url, db)

    admin_title = build_textcard_title(f"{type_name}审批", "已处理")
    admin_content = build_textcard_content(
        summary=f"你已{status_text} {applicant_name} 的{type_name}申请",
        lines=[
            ("日期", str(request.date)),
            ("时长", f"{request.hours} 小时"),
            ("审批意见", admin_comment)
        ],
        footer="点击查看详情"
    )
    send_wechat_message(approver.id, admin_title, admin_content, link_url, db)

    webhook_config = webhook_utils.load_webhook_config()
    webhook_utils.notify_time_off_approved(
        user_name=applicant_name,
        date=str(request.date),
        hours=request.hours,
        approved=approved,
        webhook_config=webhook_config,
        admin_comment=admin_comment,
        link_url=link_url
    )

    return {
        "record_type": "time_off",
        "status_text": status_text,
        "type_name": type_name,
        "applicant_name": applicant_name,
        "date": request.date,
        "hours": request.hours
    }

def approve_overtime_record_entry(
    record: database.OvertimeRecord,
    approver: database.User,
    approved: bool,
    admin_comment: Optional[str],
    db: Session
):
    record.status = "approved" if approved else "rejected"
    record.approved_by = approver.id
    record.admin_comment = admin_comment
    db.commit()

    status_text = "已确认" if approved else "已拒绝"
    link_url = "https://x.dysobo.cn/kq/?page=overtime"
    applicant = db.query(database.User).filter(database.User.id == record.user_id).first()
    applicant_name = applicant.name if applicant else "未知"

    title = build_textcard_title("加班记录", status_text)
    content = build_textcard_content(
        summary=f"您的加班申请{status_text}",
        lines=[
            ("日期", str(record.date)),
            ("时长", f"{record.hours} 小时"),
            ("审批意见", admin_comment)
        ],
        footer="点击查看详情"
    )
    send_wechat_message(record.user_id, title, content, link_url, db)

    admin_title = build_textcard_title("加班确认", "已处理")
    admin_content = build_textcard_content(
        summary=f"你已{status_text} {applicant_name} 的加班申请",
        lines=[
            ("日期", str(record.date)),
            ("时长", f"{record.hours} 小时"),
            ("审批意见", admin_comment)
        ],
        footer="点击查看详情"
    )
    send_wechat_message(approver.id, admin_title, admin_content, link_url, db)

    webhook_config = webhook_utils.load_webhook_config()
    webhook_utils.notify_overtime_approved(
        user_name=applicant_name,
        date=str(record.date),
        hours=record.hours,
        approved=approved,
        webhook_config=webhook_config,
        admin_comment=admin_comment,
        link_url=link_url
    )

    return {
        "record_type": "overtime",
        "status_text": status_text,
        "type_name": "加班",
        "applicant_name": applicant_name,
        "date": record.date,
        "hours": record.hours
    }

def get_latest_pending_approval_target(db: Session):
    latest_time_off = db.query(database.TimeOffRequest).filter(
        database.TimeOffRequest.status == "pending"
    ).order_by(
        database.TimeOffRequest.created_at.desc(),
        database.TimeOffRequest.id.desc()
    ).first()

    latest_overtime = db.query(database.OvertimeRecord).filter(
        database.OvertimeRecord.status == "pending"
    ).order_by(
        database.OvertimeRecord.created_at.desc(),
        database.OvertimeRecord.id.desc()
    ).first()

    candidates = []
    if latest_time_off:
        candidates.append(("time_off", latest_time_off, latest_time_off.created_at or datetime.min, latest_time_off.id))
    if latest_overtime:
        candidates.append(("overtime", latest_overtime, latest_overtime.created_at or datetime.min, latest_overtime.id))

    if not candidates:
        return None, None

    candidates.sort(key=lambda item: (item[2], item[3]), reverse=True)
    return candidates[0][0], candidates[0][1]

def send_wechat_command_feedback(
    wechat_user_id: str,
    success: bool,
    summary: str,
    lines: Optional[List[tuple[str, Optional[str]]]],
    db: Session,
    footer: Optional[str] = None,
    link_url: str = "https://x.dysobo.cn/kq/"
):
    if not wechat_user_id:
        return False

    title = build_textcard_title("微信指令", "已受理" if success else "处理失败")
    content = build_textcard_content(summary=summary, lines=lines or [], footer=footer)
    return send_wechat_message_to_wechat_user(wechat_user_id, title, content, link_url, db)

def handle_wechat_text_command(wechat_user_id: str, command_text: str, db: Session):
    user = db.query(database.User).filter(database.User.wechat_user_id == wechat_user_id).first()
    if not user:
        raise ValueError("未找到已绑定当前企业微信的系统用户，请先在系统中完成绑定")

    if re.match(r"^(同意|不同意|拒绝)(?:\s|$)", re.sub(r"\s+", " ", (command_text or "").strip())):
        if user.role != "admin":
            raise ValueError("只有管理员可以使用快捷审批，请按加班或调休格式发送申请")

        approval_command = parse_wechat_approval_command(command_text)
        target_type, target_record = get_latest_pending_approval_target(db)
        if not target_record:
            raise ValueError("当前没有待审批的加班或调休申请")

        if target_type == "time_off":
            approval_result = approve_time_off_request_record(
                request=target_record,
                approver=user,
                approved=approval_command["approved"],
                admin_comment=approval_command["admin_comment"],
                db=db
            )
            link_url = "https://x.dysobo.cn/kq/?page=timeoff"
        else:
            approval_result = approve_overtime_record_entry(
                record=target_record,
                approver=user,
                approved=approval_command["approved"],
                admin_comment=approval_command["admin_comment"],
                db=db
            )
            link_url = "https://x.dysobo.cn/kq/?page=overtime"

        send_wechat_command_feedback(
            wechat_user_id=wechat_user_id,
            success=True,
            summary=f"已处理最近一条待审批{approval_result['type_name']}申请",
            lines=[
                ("申请人", approval_result["applicant_name"]),
                ("日期", str(approval_result["date"])),
                ("时长", f"{approval_result['hours']} 小时"),
                ("处理结果", approval_result["status_text"]),
                ("审批意见", approval_command["admin_comment"])
            ],
            footer="审批指令格式：同意 已处理；不同意 工时不足",
            link_url=link_url,
            db=db
        )
        return {"action": "approval", "target_type": target_type}

    parsed_command = parse_wechat_command(command_text)
    if parsed_command["action"] == "overtime":
        record = create_overtime_record_for_user(
            user=user,
            record_date=parsed_command["date"],
            hours=parsed_command["hours"],
            reason=parsed_command["reason"],
            db=db
        )
        send_wechat_command_feedback(
            wechat_user_id=wechat_user_id,
            success=True,
            summary="已为你创建加班申请",
            lines=[
                ("日期", str(record.date)),
                ("时长", f"{record.hours} 小时"),
                ("事由", record.reason),
                ("状态", "待确认")
            ],
            footer="可继续发送“记加班 2h 今天 设备调试”快速登记",
            link_url="https://x.dysobo.cn/kq/?page=overtime",
            db=db
        )
        return {"action": "overtime", "id": record.id}

    request = create_time_off_request_for_user(
        user=user,
        request_date=parsed_command["date"],
        hours=parsed_command["hours"],
        reason=parsed_command["reason"],
        request_type="U",
        db=db
    )
    send_wechat_command_feedback(
        wechat_user_id=wechat_user_id,
        success=True,
        summary="已为你创建调休申请",
        lines=[
            ("日期", str(request.date)),
            ("时长", f"{request.hours} 小时"),
            ("类型", TIME_OFF_TYPE_NAMES.get(request.type, "调休")),
            ("事由", request.reason),
            ("状态", "待审批")
        ],
        footer="可继续发送“调休 1h 今天 外出办事”快速申请",
        link_url="https://x.dysobo.cn/kq/?page=timeoff",
        db=db
    )
    return {"action": "time_off", "id": request.id}


def build_shift_notification(title: str, date_value, shift_type: str, note: Optional[str] = None, highlight_text: Optional[str] = None):
    content = build_textcard_content(
        summary=highlight_text or f"你有近期 {date_value} 的排班计划",
        lines=[
            ("日期", str(date_value)),
            ("班次", shift_type),
            ("备注", note)
        ],
        footer="请注意查看"
    )
    return title, content

# --- 排班管理 ---

@app.get("/api/shifts/team")
def get_team_shifts(current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    from datetime import timedelta
    today = date.today()
    end_date = today + timedelta(days=29)
    
    shifts = db.query(database.Shift).filter(
        database.Shift.date >= today,
        database.Shift.date <= end_date
    ).order_by(database.Shift.date, database.Shift.user_id).all()
    
    result = []
    for s in shifts:
        user = db.query(database.User).filter(database.User.id == s.user_id).first()
        result.append({
            "id": s.id,
            "user_id": s.user_id,
            "user_name": user.name if user else "未知",
            "phone": user.phone if user and user.phone else "-",
            "date": str(s.date),
            "shift_type": s.shift_type,
            "note": s.note if s.note else "-"
        })
    return result

@app.get("/api/shifts")
def list_shifts(start_date: Optional[date] = None, end_date: Optional[date] = None, user_id: Optional[int] = None, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        if user_id is not None and user_id != current_user.id:
            raise HTTPException(status_code=403, detail="无权限查看他人排班")
        user_id = current_user.id

    query = db.query(database.Shift)
    if start_date:
        query = query.filter(database.Shift.date >= start_date)
    if end_date:
        query = query.filter(database.Shift.date <= end_date)
    if user_id:
        query = query.filter(database.Shift.user_id == user_id)
    
    shifts = query.order_by(database.Shift.date).all()
    result = []
    for s in shifts:
        user = db.query(database.User).filter(database.User.id == s.user_id).first()
        result.append({
            "id": s.id,
            "user_id": s.user_id,
            "user_name": user.name if user else "未知",
            "phone": user.phone if user and user.phone else "-",
            "date": str(s.date),
            "shift_type": s.shift_type,
            "note": s.note
        })
    return result

@app.post("/api/shifts")
def create_shift(shift_data: ShiftCreate, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    user_id = int(shift_data.user_id) if isinstance(shift_data.user_id, str) else shift_data.user_id
    
    existing = db.query(database.Shift).filter(
        database.Shift.user_id == user_id,
        database.Shift.date == shift_data.date
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="该日期已存在排班")
    
    shift = database.Shift(user_id=user_id, date=shift_data.date, shift_type=shift_data.shift_type, note=shift_data.note)
    db.add(shift)
    db.commit()
    db.refresh(shift)

    shift_title, shift_content = build_shift_notification("排班通知 | 最新安排", shift_data.date, shift_data.shift_type, shift_data.note)
    send_wechat_message(user_id, shift_title, shift_content, "https://x.dysobo.cn/kq/", db)

    return {"id": shift.id, "message": "排班创建成功"}

@app.put("/api/shifts/{shift_id}")
def update_shift(shift_id: int, shift_data: ShiftUpdate, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    shift = db.query(database.Shift).filter(database.Shift.id == shift_id).first()
    if not shift:
        raise HTTPException(status_code=404, detail="排班不存在")
    
    if shift_data.shift_type is not None:
        shift.shift_type = shift_data.shift_type
    if shift_data.note is not None:
        shift.note = shift_data.note

    db.commit()
    db.refresh(shift)

    shift_title, shift_content = build_shift_notification(
        "排班通知 | 排班更新",
        shift.date,
        shift.shift_type,
        shift.note,
        f"你的 {shift.date} 排班已更新"
    )
    send_wechat_message(shift.user_id, shift_title, shift_content, "https://x.dysobo.cn/kq/", db)

    return {"id": shift.id, "message": "排班更新成功"}

@app.delete("/api/shifts/{shift_id}")
def delete_shift(shift_id: int, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    shift = db.query(database.Shift).filter(database.Shift.id == shift_id).first()
    if not shift:
        raise HTTPException(status_code=404, detail="排班不存在")
    
    db.delete(shift)
    db.commit()
    return {"message": "排班删除成功"}

@app.post("/api/shifts/{shift_id}/notify")
def notify_shift(shift_id: int, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    """管理员手动推送排班提醒"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")

    shift = db.query(database.Shift).filter(database.Shift.id == shift_id).first()
    if not shift:
        raise HTTPException(status_code=404, detail="排班不存在")

    title, content = build_shift_notification("产品试验室提醒 | 排班提醒", shift.date, shift.shift_type, shift.note)
    success = send_wechat_message(shift.user_id, title, content, "https://x.dysobo.cn/kq/", db)

    if success:
        return {"message": "排班提醒已发送"}
    else:
        raise HTTPException(status_code=500, detail="推送失败，请检查该用户是否绑定企业微信")

# --- 调休申请 ---

@app.get("/api/time-off")
def list_time_off(user_id: Optional[int] = None, status: Optional[str] = None, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        if user_id is not None and user_id != current_user.id:
            raise HTTPException(status_code=403, detail="无权限查看他人申请")
        user_id = current_user.id

    query = db.query(database.TimeOffRequest)
    if user_id:
        query = query.filter(database.TimeOffRequest.user_id == user_id)
    if status:
        query = query.filter(database.TimeOffRequest.status == status)
    
    requests_list = query.order_by(
        database.TimeOffRequest.date.desc(),
        database.TimeOffRequest.created_at.desc(),
        database.TimeOffRequest.id.desc()
    ).all()
    result = []
    for r in requests_list:
        user = db.query(database.User).filter(database.User.id == r.user_id).first()
        result.append({
            "id": r.id,
            "user_id": r.user_id,
            "user_name": user.name if user else "未知",
            "date": str(r.date),
            "hours": r.hours,
            "type": r.type,
            "reason": r.reason,
            "admin_comment": r.admin_comment,
            "status": r.status,
            "created_at": str(r.created_at)
        })
    return result

@app.post("/api/time-off")
def create_time_off(request_data: TimeOffRequestCreate, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    request = create_time_off_request_for_user(
        user=current_user,
        request_date=request_data.date,
        hours=request_data.hours,
        reason=request_data.reason,
        request_type=request_data.type,
        db=db
    )
    return {"id": request.id, "message": "调休申请已提交"}

@app.post("/api/time-off/{request_id}/approve")
def approve_time_off(request_id: int, approve_data: TimeOffRequestApprove, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    request = db.query(database.TimeOffRequest).filter(database.TimeOffRequest.id == request_id).first()
    if not request:
        raise HTTPException(status_code=404, detail="申请不存在")

    approve_time_off_request_record(
        request=request,
        approver=current_user,
        approved=approve_data.approved,
        admin_comment=approve_data.admin_comment,
        db=db
    )

    return {"message": "申请已" + ("批准" if approve_data.approved else "拒绝")}

@app.put("/api/time-off/{request_id}")
def update_time_off(request_id: int, update_data: TimeOffRequestCreate, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    request = db.query(database.TimeOffRequest).filter(database.TimeOffRequest.id == request_id).first()
    if not request:
        raise HTTPException(status_code=404, detail="申请不存在")
    
    request.date = update_data.date
    request.hours = update_data.hours
    request.type = update_data.type
    request.reason = update_data.reason
    request.updated_at = datetime.now()
    db.commit()
    db.refresh(request)
    return {"message": "申请已更新", "id": request.id}

@app.delete("/api/time-off/{request_id}")
def delete_time_off(request_id: int, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    request = db.query(database.TimeOffRequest).filter(database.TimeOffRequest.id == request_id).first()
    if not request:
        raise HTTPException(status_code=404, detail="申请不存在")
    
    if current_user.role != "admin":
        if request.user_id != current_user.id:
            raise HTTPException(status_code=403, detail="无权限删除")
        if request.status not in ["pending", "rejected"]:
            raise HTTPException(status_code=400, detail="已批准，不可删除")
    
    db.delete(request)
    db.commit()
    return {"message": "申请已删除"}

# --- 加班记录 ---

@app.get("/api/overtime")
def list_overtime(user_id: Optional[int] = None, status: Optional[str] = None, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        if user_id is not None and user_id != current_user.id:
            raise HTTPException(status_code=403, detail="无权限查看他人记录")
        user_id = current_user.id

    query = db.query(database.OvertimeRecord)
    if user_id:
        query = query.filter(database.OvertimeRecord.user_id == user_id)
    if status:
        query = query.filter(database.OvertimeRecord.status == status)
    
    records = query.order_by(
        database.OvertimeRecord.date.desc(),
        database.OvertimeRecord.created_at.desc(),
        database.OvertimeRecord.id.desc()
    ).all()
    result = []
    for r in records:
        user = db.query(database.User).filter(database.User.id == r.user_id).first()
        result.append({
            "id": r.id,
            "user_id": r.user_id,
            "user_name": user.name if user else "未知",
            "date": str(r.date),
            "hours": r.hours,
            "reason": r.reason,
            "admin_comment": r.admin_comment,
            "status": r.status,
            "created_at": str(r.created_at)
        })
    return result

@app.get("/api/webhook/config")
def get_webhook_config(current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        return webhook_utils.default_webhook_config()
    return webhook_utils.load_webhook_config()

@app.post("/api/webhook/config")
def update_webhook_config(webhook_config: WebhookConfig, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    config = webhook_config.dict()
    webhook_utils.save_webhook_config(config)
    return {"message": "配置已保存", "config": config}

@app.post("/api/webhook/test")
def test_webhook(webhook_test: dict, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    config = webhook_utils.load_webhook_config()
    title = webhook_test.get("title", "测试推送")
    content = webhook_test.get("content", "这是一条测试消息")
    success = webhook_utils.send_webhook(config, title, content)
    if success:
        return {"message": "测试推送已发送"}
    else:
        raise HTTPException(status_code=500, detail="推送失败，请检查配置")

@app.post("/api/overtime")
def create_overtime(record_data: OvertimeRecordCreate, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    record = create_overtime_record_for_user(
        user=current_user,
        record_date=record_data.date,
        hours=record_data.hours,
        reason=record_data.reason,
        db=db
    )
    return {"id": record.id, "message": "加班记录已提交"}

@app.post("/api/overtime/{record_id}/approve")
def approve_overtime(record_id: int, approve_data: OvertimeRecordApprove, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    record = db.query(database.OvertimeRecord).filter(database.OvertimeRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")

    approve_overtime_record_entry(
        record=record,
        approver=current_user,
        approved=approve_data.approved,
        admin_comment=approve_data.admin_comment,
        db=db
    )

    return {"message": "加班记录已" + ("批准" if approve_data.approved else "拒绝")}

@app.put("/api/overtime/{record_id}")
def update_overtime(record_id: int, update_data: OvertimeRecordCreate, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    record = db.query(database.OvertimeRecord).filter(database.OvertimeRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    record.date = update_data.date
    record.hours = update_data.hours
    record.reason = update_data.reason
    db.commit()
    db.refresh(record)
    return {"message": "记录已更新", "id": record.id}

@app.delete("/api/overtime/{record_id}")
def delete_overtime(record_id: int, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    record = db.query(database.OvertimeRecord).filter(database.OvertimeRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    if current_user.role != "admin":
        if record.user_id != current_user.id:
            raise HTTPException(status_code=403, detail="无权限删除")
        if record.status not in ["pending", "rejected"]:
            raise HTTPException(status_code=400, detail="已确认，不可删除")
    
    db.delete(record)
    db.commit()
    return {"message": "记录已删除"}

# --- 统计汇总 ---

@app.get("/api/stats/summary")
def get_summary(current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    from datetime import datetime
    current_year = datetime.now().year
    year_start = date(current_year, 1, 1)
    
    if current_user.role == "admin":
        users = db.query(database.User).filter(database.User.role == "member").all()
    else:
        users = [current_user]
    
    total_approved_time_off = 0
    total_pending_time_off = 0
    total_overtime_hours = 0.0
    total_pending_overtime = 0
    
    for user in users:
        approved_time_off_records = db.query(database.TimeOffRequest).filter(
            database.TimeOffRequest.user_id == user.id,
            database.TimeOffRequest.status == "approved",
            database.TimeOffRequest.date >= year_start
        ).all()
        approved_time_off = sum(r.hours for r in approved_time_off_records)
        
        pending_time_off_records = db.query(database.TimeOffRequest).filter(
            database.TimeOffRequest.user_id == user.id,
            database.TimeOffRequest.status == "pending",
            database.TimeOffRequest.date >= year_start
        ).all()
        pending_time_off = sum(r.hours for r in pending_time_off_records)
        
        approved_overtime = db.query(database.OvertimeRecord).filter(
            database.OvertimeRecord.user_id == user.id,
            database.OvertimeRecord.status == "approved",
            database.OvertimeRecord.date >= year_start
        ).all()
        total_overtime_hours += sum(r.hours for r in approved_overtime)
        
        pending_overtime = db.query(database.OvertimeRecord).filter(
            database.OvertimeRecord.user_id == user.id,
            database.OvertimeRecord.status == "pending"
        ).count()
        
        total_approved_time_off += approved_time_off
        total_pending_time_off += pending_time_off
        total_pending_overtime += pending_overtime
    
    return {
        "is_admin": current_user.role == "admin",
        "total_users": len(users),
        "time_off_approved": total_approved_time_off,
        "time_off_pending": total_pending_time_off,
        "overtime_hours": total_overtime_hours,
        "overtime_pending": total_pending_overtime
    }

@app.get("/api/stats/my-summary")
def get_my_summary(current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    approved_time_off = db.query(database.TimeOffRequest).filter(
        database.TimeOffRequest.user_id == current_user.id,
        database.TimeOffRequest.status == "approved"
    ).count()
    pending_time_off = db.query(database.TimeOffRequest).filter(
        database.TimeOffRequest.user_id == current_user.id,
        database.TimeOffRequest.status == "pending"
    ).count()
    
    approved_overtime = db.query(database.OvertimeRecord).filter(
        database.OvertimeRecord.user_id == current_user.id,
        database.OvertimeRecord.status == "approved"
    ).all()
    total_overtime_hours = sum(r.hours for r in approved_overtime)
    pending_overtime = db.query(database.OvertimeRecord).filter(
        database.OvertimeRecord.user_id == current_user.id,
        database.OvertimeRecord.status == "pending"
    ).count()
    
    return {
        "user_name": current_user.name,
        "time_off_approved": approved_time_off,
        "time_off_pending": pending_time_off,
        "overtime_hours": total_overtime_hours,
        "overtime_pending": pending_overtime
    }

# ==================== 数据导出与备份 ====================

def normalize_export_hour(value: float) -> Union[float, int]:
    normalized = round(float(value or 0), 1)
    return int(normalized) if normalized.is_integer() else normalized

def get_export_month_bounds(year: int, month: int) -> tuple[date, date]:
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="月份必须在 1-12 之间")
    month_start = date(year, month, 1)
    month_end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return month_start, month_end

def count_month_workdays(month_start: date, month_end: date) -> int:
    workdays = 0
    cursor = month_start
    while cursor < month_end:
        if cursor.weekday() < 5:
            workdays += 1
        cursor += timedelta(days=1)
    return workdays

def autosize_worksheet_columns(worksheet):
    from openpyxl.utils import get_column_letter

    for column_index, column_cells in enumerate(worksheet.iter_cols(min_col=1, max_col=worksheet.max_column, min_row=1, max_row=worksheet.max_row), start=1):
        max_length = 0
        column_letter = get_column_letter(column_index)
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 24)

def resolve_monthly_export_users(db: Session, user_ids: List[int]) -> List[database.User]:
    members = db.query(database.User).filter(database.User.role == "member").order_by(database.User.id).all()
    member_map = {user.id: user for user in members}

    if not user_ids:
        if not members:
            raise HTTPException(status_code=400, detail="当前没有可导出的组员")
        return members

    resolved_users = []
    seen_ids = set()
    for raw_user_id in user_ids:
        try:
            user_id = int(raw_user_id)
        except (TypeError, ValueError):
            continue
        if user_id in member_map and user_id not in seen_ids:
            resolved_users.append(member_map[user_id])
            seen_ids.add(user_id)

    if not resolved_users:
        raise HTTPException(status_code=400, detail="请选择至少一名组员用于导出")

    return resolved_users

def build_monthly_export_workbook(year: int, month: int, ordered_user_ids: List[int], db: Session):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，请先安装后再导出")

    users = resolve_monthly_export_users(db, ordered_user_ids)
    month_start, month_end = get_export_month_bounds(year, month)
    _, days_in_month = calendar.monthrange(year, month)
    workday_hours = count_month_workdays(month_start, month_end) * 8
    user_ids = [user.id for user in users]

    time_off_records = db.query(database.TimeOffRequest).filter(
        database.TimeOffRequest.user_id.in_(user_ids),
        database.TimeOffRequest.date >= month_start,
        database.TimeOffRequest.date < month_end
    ).order_by(
        database.TimeOffRequest.user_id,
        database.TimeOffRequest.date,
        database.TimeOffRequest.id
    ).all()

    overtime_records = db.query(database.OvertimeRecord).filter(
        database.OvertimeRecord.user_id.in_(user_ids),
        database.OvertimeRecord.date >= month_start,
        database.OvertimeRecord.date < month_end
    ).order_by(
        database.OvertimeRecord.user_id,
        database.OvertimeRecord.date,
        database.OvertimeRecord.id
    ).all()

    time_off_by_user = defaultdict(list)
    approved_time_off_by_user = defaultdict(list)
    for record in time_off_records:
        time_off_by_user[record.user_id].append(record)
        if record.status == "approved":
            approved_time_off_by_user[record.user_id].append(record)

    overtime_by_user = defaultdict(list)
    approved_overtime_by_user = defaultdict(list)
    for record in overtime_records:
        overtime_by_user[record.user_id].append(record)
        if record.status == "approved":
            approved_overtime_by_user[record.user_id].append(record)

    approved_time_off_hours_by_user_day = defaultdict(float)
    approved_overtime_hours_by_user_day = defaultdict(float)
    for record in time_off_records:
        if record.status == "approved":
            approved_time_off_hours_by_user_day[(record.user_id, record.date)] += float(record.hours or 0)
    for record in overtime_records:
        if record.status == "approved":
            approved_overtime_hours_by_user_day[(record.user_id, record.date)] += float(record.hours or 0)

    workbook = Workbook()
    worksheet_raw = workbook.active
    worksheet_raw.title = "原始记录"
    worksheet_stats = workbook.create_sheet("考勤统计")
    worksheet_summary = workbook.create_sheet("数据统计")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="EAF3FF")
    thin_side = Side(style="thin", color="D6DCE5")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    # Sheet1: 原始记录，保持旧导出结构
    raw_header = ["姓名"] + [f"{day}日" for day in range(1, days_in_month + 1)]
    worksheet_raw.append(raw_header)
    for cell in worksheet_raw[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment

    for user in users:
        daily_data = defaultdict(list)
        for record in time_off_by_user.get(user.id, []):
            daily_data[record.date.day].append(f"{record.type or 'U'}{normalize_export_hour(record.hours)}")
        for record in overtime_by_user.get(user.id, []):
            daily_data[record.date.day].append(f"▲{normalize_export_hour(record.hours)}")

        row = [user.name]
        for day in range(1, days_in_month + 1):
            row.append(" ".join(daily_data.get(day, [])))
        worksheet_raw.append(row)

    for row in worksheet_raw.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment if cell.column != 1 else left_alignment
    worksheet_raw.freeze_panes = "B2"

    # Sheet2: 考勤统计
    stats_title = f"产品试验室{year}年{month}月考勤统计"
    worksheet_stats.merge_cells("A1:W1")
    worksheet_stats["A1"] = stats_title
    worksheet_stats["A1"].font = Font(bold=True, size=14)
    worksheet_stats["A1"].alignment = center_alignment
    worksheet_stats["A1"].fill = title_fill
    worksheet_stats["A1"].border = thin_border

    stats_headers = [
        "人员编号", "姓名", "组别", "应出勤（h）", "实出勤（h）", "迟到早退（次）",
        "事假（h）", "病假（h）", "工伤假（h）", "婚假（h）", "产护理假（h）",
        "经期假（h）", "孕假（h）", "哺乳假（h）", "探亲假（h）", "年休假（h）",
        "旷工（h）", "丧假（h）", "加班（h）", "调休（h）", "外出公干（h）",
        "其他", "签字确认"
    ]
    worksheet_stats.append([])
    worksheet_stats.append(stats_headers)
    for cell in worksheet_stats[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment

    for index, user in enumerate(users, start=1):
        time_off_totals = {
            "S": 0.0,
            "B": 0.0,
            "H": 0.0,
            "C_L": 0.0,
            "J": 0.0,
            "Y": 0.0,
            "R": 0.0,
            "T": 0.0,
            "N": 0.0,
            "Z": 0.0,
            "U": 0.0
        }
        overtime_hours = 0.0

        for record in approved_time_off_by_user.get(user.id, []):
            if record.type == "U":
                time_off_totals["U"] += float(record.hours or 0)
            elif record.type == "B":
                time_off_totals["B"] += float(record.hours or 0)
            elif record.type == "S":
                time_off_totals["S"] += float(record.hours or 0)
            elif record.type == "H":
                time_off_totals["H"] += float(record.hours or 0)
            elif record.type in ["C", "L"]:
                time_off_totals["C_L"] += float(record.hours or 0)
            elif record.type == "J":
                time_off_totals["J"] += float(record.hours or 0)
            elif record.type == "Y":
                time_off_totals["Y"] += float(record.hours or 0)
            elif record.type == "R":
                time_off_totals["R"] += float(record.hours or 0)
            elif record.type == "T":
                time_off_totals["T"] += float(record.hours or 0)
            elif record.type == "N":
                time_off_totals["N"] += float(record.hours or 0)
            elif record.type == "Z":
                time_off_totals["Z"] += float(record.hours or 0)

        for record in approved_overtime_by_user.get(user.id, []):
            overtime_hours += float(record.hours or 0)

        row_index = index + 3
        worksheet_stats.append([
            index,
            user.name,
            "试验",
            normalize_export_hour(workday_hours),
            None,
            0,
            normalize_export_hour(time_off_totals["S"]),
            normalize_export_hour(time_off_totals["B"]),
            0,
            normalize_export_hour(time_off_totals["H"]),
            normalize_export_hour(time_off_totals["C_L"]),
            normalize_export_hour(time_off_totals["J"]),
            normalize_export_hour(time_off_totals["Y"]),
            normalize_export_hour(time_off_totals["R"]),
            normalize_export_hour(time_off_totals["T"]),
            normalize_export_hour(time_off_totals["N"]),
            0,
            normalize_export_hour(time_off_totals["Z"]),
            normalize_export_hour(overtime_hours),
            normalize_export_hour(time_off_totals["U"]),
            0,
            0,
            ""
        ])

        worksheet_stats[f"E{row_index}"] = f"=D{row_index}-SUM(G{row_index}:R{row_index})+S{row_index}-T{row_index}-U{row_index}-V{row_index}"

    for row in worksheet_stats.iter_rows(min_row=3, max_row=worksheet_stats.max_row, min_col=1, max_col=23):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    for row_index in range(4, worksheet_stats.max_row + 1):
        for column_letter in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V"]:
            cell = worksheet_stats[f"{column_letter}{row_index}"]
            cell.number_format = "General"
            if cell.value == 0:
                cell.font = Font(color="FFFFFF")

    export_time_row = worksheet_stats.max_row + 2
    worksheet_stats.merge_cells(start_row=export_time_row, start_column=22, end_row=export_time_row, end_column=23)
    worksheet_stats.cell(row=export_time_row, column=22).value = f"填写时间：{datetime.now().year}年{datetime.now().month}月{datetime.now().day}日"
    worksheet_stats.cell(row=export_time_row, column=22).alignment = Alignment(horizontal="right", vertical="center")
    worksheet_stats.cell(row=export_time_row, column=22).font = Font(italic=True)

    # Sheet3: 数据统计
    summary_title = f"产品试验室{year}年{month}月数据统计"
    worksheet_summary.merge_cells("A1:H1")
    worksheet_summary["A1"] = summary_title
    worksheet_summary["A1"].font = Font(bold=True, size=14)
    worksheet_summary["A1"].alignment = center_alignment
    worksheet_summary["A1"].fill = title_fill
    worksheet_summary["A1"].border = thin_border
    worksheet_summary.append([])
    worksheet_summary.append(["人员编号", "姓名", "正餐合计", "大夜班", "小夜班", "交通费（天）", "调休（h）", "加班（h）"])

    for cell in worksheet_summary[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment

    for index, user in enumerate(users, start=1):
        stats_row = index + 3
        meal_count = 0
        large_night_count = 0
        small_night_count = 0
        transport_days = 0

        current_day = month_start
        while current_day < month_end:
            leave_hours = approved_time_off_hours_by_user_day[(user.id, current_day)]
            overtime_hours = approved_overtime_hours_by_user_day[(user.id, current_day)]
            is_workday = current_day.weekday() < 5

            if is_workday:
                if leave_hours < 8:
                    meal_count += 1
                    transport_days += 1
                if overtime_hours > 1:
                    meal_count += 1
                if overtime_hours >= 9:
                    large_night_count += 1
                if overtime_hours >= 5:
                    small_night_count += 1
            else:
                if overtime_hours > 0:
                    meal_count += 1
                    transport_days += 1
                    if overtime_hours > 9:
                        meal_count += 1
                if overtime_hours >= 16:
                    large_night_count += 1
                if overtime_hours >= 12:
                    small_night_count += 1

            current_day += timedelta(days=1)

        worksheet_summary.append([
            index,
            user.name,
            meal_count,
            large_night_count,
            small_night_count,
            transport_days,
            f"='考勤统计'!T{stats_row}",
            f"='考勤统计'!S{stats_row}"
        ])

    for row in worksheet_summary.iter_rows(min_row=3, max_row=worksheet_summary.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    autosize_worksheet_columns(worksheet_raw)
    autosize_worksheet_columns(worksheet_stats)
    autosize_worksheet_columns(worksheet_summary)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue(), f"考勤统计_{year}年{month}月.xlsx"

@app.get("/api/export/monthly")
def export_monthly_stats(month: Optional[int] = None, year: Optional[int] = None, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")

    if not month or not year:
        today = date.today()
        month = today.month
        year = today.year

    file_bytes, filename = build_monthly_export_workbook(year, month, [], db)
    encoded_filename = quote(filename)
    return Response(
        content=file_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

@app.post("/api/export/monthly")
def export_monthly_stats_with_order(export_request: MonthlyExportRequest, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")

    month = export_request.month or date.today().month
    year = export_request.year or date.today().year
    file_bytes, filename = build_monthly_export_workbook(year, month, export_request.user_ids, db)
    encoded_filename = quote(filename)
    return Response(
        content=file_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

@app.get("/api/backup/export")
def export_backup(include_sensitive: bool = False, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    users = db.query(database.User).all()
    shifts = db.query(database.Shift).all()
    time_off = db.query(database.TimeOffRequest).all()
    overtime = db.query(database.OvertimeRecord).all()
    
    wechat_config = db.query(database.WechatConfig).first()

    backup_data = {
        "export_time": datetime.now().isoformat(),
        "version": "2.0",
        "includes_sensitive_data": include_sensitive,
        "users": [
            {
                "id": u.id, "name": u.name, "password": u.password if include_sensitive else None, "role": u.role,
                "phone": u.phone, "wechat_user_id": u.wechat_user_id,
                "enable_push": u.enable_push, "created_at": str(u.created_at)
            } for u in users
        ],
        "shifts": [
            {
                "id": s.id, "user_id": s.user_id, "date": str(s.date),
                "shift_type": s.shift_type, "note": s.note, "created_at": str(s.created_at)
            } for s in shifts
        ],
        "time_off_requests": [
            {
                "id": t.id, "user_id": t.user_id, "date": str(t.date),
                "hours": t.hours, "type": t.type, "reason": t.reason, "status": t.status,
                "approved_by": t.approved_by, "admin_comment": t.admin_comment, "created_at": str(t.created_at),
                "updated_at": str(t.updated_at)
            } for t in time_off
        ],
        "overtime_records": [
            {
                "id": o.id, "user_id": o.user_id, "date": str(o.date),
                "hours": o.hours, "reason": o.reason, "status": o.status,
                "approved_by": o.approved_by, "admin_comment": o.admin_comment, "created_at": str(o.created_at)
            } for o in overtime
        ],
        "wechat_config": {
            "api_url": wechat_config.api_url,
            "corp_id": wechat_config.corp_id,
            "secret": wechat_config.secret if include_sensitive else "",
            "agent_id": wechat_config.agent_id,
            "token": wechat_config.token if include_sensitive else "",
            "encoding_aes_key": wechat_config.encoding_aes_key if include_sensitive else "",
            "enabled": wechat_config.enabled,
            "updated_at": str(wechat_config.updated_at)
        } if wechat_config else None
    }
    
    return {"filename": f"考勤备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json", "data": backup_data}

@app.post("/api/backup/import")
def import_backup(backup_data: dict, current_user: database.User = Depends(get_current_user), db: Session = Depends(database.get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    try:
        if "users" not in backup_data:
            raise HTTPException(status_code=400, detail="无效的备份文件格式")

        imported_users = backup_data.get("users", [])
        imported_user_ids = {u["id"] for u in imported_users}

        for s in backup_data.get("shifts", []):
            if s["user_id"] not in imported_user_ids:
                raise HTTPException(status_code=400, detail=f"排班数据引用了不存在的用户：{s['user_id']}")

        for t in backup_data.get("time_off_requests", []):
            if t["user_id"] not in imported_user_ids:
                raise HTTPException(status_code=400, detail=f"调休数据引用了不存在的用户：{t['user_id']}")
            if t.get("approved_by") is not None and t["approved_by"] not in imported_user_ids:
                raise HTTPException(status_code=400, detail=f"调休审批人不存在：{t['approved_by']}")

        for o in backup_data.get("overtime_records", []):
            if o["user_id"] not in imported_user_ids:
                raise HTTPException(status_code=400, detail=f"加班数据引用了不存在的用户：{o['user_id']}")
            if o.get("approved_by") is not None and o["approved_by"] not in imported_user_ids:
                raise HTTPException(status_code=400, detail=f"加班审批人不存在：{o['approved_by']}")

        db.query(database.OvertimeRecord).delete()
        db.query(database.TimeOffRequest).delete()
        db.query(database.Shift).delete()
        db.query(database.WechatConfig).delete()
        db.query(database.User).delete()

        for u in backup_data.get("users", []):
            user = database.User(
                id=u["id"], name=u["name"], role=u["role"],
                phone=u.get("phone"),
                password=u.get("password", get_password_hash("123456")),
                wechat_user_id=u.get("wechat_user_id"),
                enable_push=u.get("enable_push", True),
                created_at=parse_datetime_value(u.get("created_at")) or datetime.now()
            )
            db.add(user)

        for s in backup_data.get("shifts", []):
            shift = database.Shift(
                id=s["id"], user_id=s["user_id"], date=parse_date_value(s["date"]),
                shift_type=s["shift_type"], note=s.get("note"),
                created_at=parse_datetime_value(s.get("created_at")) or datetime.now()
            )
            db.add(shift)

        for t in backup_data.get("time_off_requests", []):
            time_off_req = database.TimeOffRequest(
                id=t["id"], user_id=t["user_id"], date=parse_date_value(t["date"]),
                hours=t["hours"], type=t.get("type", "U"), reason=t.get("reason"),
                status=t["status"], approved_by=t.get("approved_by"), admin_comment=t.get("admin_comment"),
                created_at=parse_datetime_value(t.get("created_at")) or datetime.now(),
                updated_at=parse_datetime_value(t.get("updated_at")) or datetime.now()
            )
            db.add(time_off_req)

        for o in backup_data.get("overtime_records", []):
            overtime_rec = database.OvertimeRecord(
                id=o["id"], user_id=o["user_id"], date=parse_date_value(o["date"]),
                hours=o["hours"], reason=o.get("reason"), status=o["status"],
                approved_by=o.get("approved_by"), admin_comment=o.get("admin_comment"),
                created_at=parse_datetime_value(o.get("created_at")) or datetime.now()
            )
            db.add(overtime_rec)

        # 恢复企业微信配置
        wc = backup_data.get("wechat_config")
        if wc:
            wechat_cfg = database.WechatConfig(
                api_url=wc.get("api_url", ""),
                corp_id=wc.get("corp_id", ""),
                secret=wc.get("secret", ""),
                agent_id=wc.get("agent_id", 0),
                token=wc.get("token", ""),
                encoding_aes_key=wc.get("encoding_aes_key", ""),
                enabled=wc.get("enabled", False),
                updated_at=parse_datetime_value(wc.get("updated_at")) or datetime.now()
            )
            db.add(wechat_cfg)
        
        db.commit()
        
        return {"message": "数据恢复成功", "count": {
            "users": len(backup_data.get("users", [])),
            "shifts": len(backup_data.get("shifts", [])),
            "time_off": len(backup_data.get("time_off_requests", [])),
            "overtime": len(backup_data.get("overtime_records", []))
        }}
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"恢复失败：{str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
