import os
import tempfile
import unittest
import io

try:
    from fastapi.testclient import TestClient
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from openpyxl import load_workbook

    import database
    import main
    import webhook as webhook_utils
except ModuleNotFoundError:
    TestClient = None
    create_engine = None
    sessionmaker = None
    load_workbook = None
    database = None
    main = None
    webhook_utils = None


@unittest.skipUnless(TestClient is not None and load_workbook is not None, "FastAPI 或 openpyxl 测试依赖未安装")
class APISecurityTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.db_path = os.path.join(cls.temp_dir.name, "attendance-test.db")
        cls.webhook_path = os.path.join(cls.temp_dir.name, "webhook-config.json")

        database.engine = create_engine(
            f"sqlite:///{cls.db_path}",
            connect_args={"check_same_thread": False}
        )
        database.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=database.engine)
        webhook_utils.WEBHOOK_CONFIG_FILE = cls.webhook_path

        database.Base.metadata.create_all(bind=database.engine)
        cls.client_cm = TestClient(main.app)
        cls.client = cls.client_cm.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.client_cm.__exit__(None, None, None)
        cls.temp_dir.cleanup()

    def setUp(self):
        database.Base.metadata.drop_all(bind=database.engine)
        database.Base.metadata.create_all(bind=database.engine)
        if os.path.exists(self.webhook_path):
            os.remove(self.webhook_path)
        main.startup()

    def login(self, name: str, password: str) -> str:
        response = self.client.post("/api/login", json={"name": name, "password": password})
        self.assertEqual(response.status_code, 200, response.text)
        return response.json()["token"]

    def admin_headers(self) -> dict:
        return {"Authorization": f"Bearer {self.login('admin', 'admin123')}"}

    def create_member(self, name: str, password: str = "member123") -> int:
        response = self.client.post(
            "/api/users",
            json={"name": name, "password": password, "role": "member"},
            headers=self.admin_headers()
        )
        self.assertEqual(response.status_code, 200, response.text)
        return response.json()["id"]

    def test_protected_lists_require_authentication(self):
        for endpoint in ["/api/shifts", "/api/time-off", "/api/overtime", "/api/shifts/team"]:
            response = self.client.get(endpoint)
            self.assertEqual(response.status_code, 403, f"{endpoint}: {response.text}")

    def test_member_cannot_list_users_but_can_read_self(self):
        self.create_member("member_a")
        member_token = self.login("member_a", "member123")
        headers = {"Authorization": f"Bearer {member_token}"}

        users_response = self.client.get("/api/users", headers=headers)
        self.assertEqual(users_response.status_code, 403, users_response.text)

        me_response = self.client.get("/api/me", headers=headers)
        self.assertEqual(me_response.status_code, 200, me_response.text)
        self.assertEqual(me_response.json()["name"], "member_a")

    def test_member_cannot_query_other_members_records(self):
        member_a_id = self.create_member("member_a")
        member_b_id = self.create_member("member_b")
        member_token = self.login("member_a", "member123")
        headers = {"Authorization": f"Bearer {member_token}"}

        response = self.client.get(f"/api/time-off?user_id={member_b_id}", headers=headers)
        self.assertEqual(response.status_code, 403, response.text)

        own_response = self.client.get(f"/api/time-off?user_id={member_a_id}", headers=headers)
        self.assertEqual(own_response.status_code, 200, own_response.text)

    def test_backup_export_masks_sensitive_data_by_default(self):
        self.create_member("member_a")
        response = self.client.get("/api/backup/export", headers=self.admin_headers())
        self.assertEqual(response.status_code, 200, response.text)

        exported_users = response.json()["data"]["users"]
        self.assertTrue(all(user["password"] is None for user in exported_users))

        sensitive_response = self.client.get(
            "/api/backup/export?include_sensitive=true",
            headers=self.admin_headers()
        )
        self.assertEqual(sensitive_response.status_code, 200, sensitive_response.text)
        sensitive_users = sensitive_response.json()["data"]["users"]
        self.assertTrue(all(user["password"] for user in sensitive_users))

    def test_invalid_backup_import_returns_400_and_rolls_back(self):
        payload = {
            "users": [
                {"id": 1, "name": "admin", "password": "hashed", "role": "admin"}
            ],
            "shifts": [
                {"id": 1, "user_id": 999, "date": "2026-04-01", "shift_type": "早班"}
            ],
            "time_off_requests": [],
            "overtime_records": []
        }

        response = self.client.post("/api/backup/import", json=payload, headers=self.admin_headers())
        self.assertEqual(response.status_code, 400, response.text)

        with database.SessionLocal() as db:
            users = db.query(database.User).all()
            self.assertEqual(len(users), 1)
            self.assertEqual(users[0].name, "admin")

    def test_admin_time_off_list_is_sorted_by_request_date_desc(self):
        self.create_member("member_a")
        member_headers = {"Authorization": f"Bearer {self.login('member_a', 'member123')}"}

        for request_date in ["2026-04-01", "2026-04-03", "2026-04-02"]:
            response = self.client.post(
                "/api/time-off",
                json={"date": request_date, "hours": 4, "type": "U", "reason": f"reason-{request_date}"},
                headers=member_headers
            )
            self.assertEqual(response.status_code, 200, response.text)

        response = self.client.get("/api/time-off", headers=self.admin_headers())
        self.assertEqual(response.status_code, 200, response.text)
        dates = [item["date"] for item in response.json()]
        self.assertEqual(dates[:3], ["2026-04-03", "2026-04-02", "2026-04-01"])

    def test_admin_overtime_list_is_sorted_by_request_date_desc(self):
        self.create_member("member_a")
        member_headers = {"Authorization": f"Bearer {self.login('member_a', 'member123')}"}

        for request_date in ["2026-04-01", "2026-04-03", "2026-04-02"]:
            response = self.client.post(
                "/api/overtime",
                json={"date": request_date, "hours": 2, "reason": f"reason-{request_date}"},
                headers=member_headers
            )
            self.assertEqual(response.status_code, 200, response.text)

        response = self.client.get("/api/overtime", headers=self.admin_headers())
        self.assertEqual(response.status_code, 200, response.text)
        dates = [item["date"] for item in response.json()]
        self.assertEqual(dates[:3], ["2026-04-03", "2026-04-02", "2026-04-01"])

    def test_parse_wechat_command_supports_relative_date_and_reason(self):
        parsed = main.parse_wechat_command("记加班 4h 今天 设备调试", today=main.date(2026, 4, 7))
        self.assertEqual(parsed["action"], "overtime")
        self.assertEqual(str(parsed["date"]), "2026-04-07")
        self.assertEqual(parsed["hours"], 4.0)
        self.assertEqual(parsed["reason"], "设备调试")

    def test_parse_wechat_command_supports_yesterday(self):
        parsed = main.parse_wechat_command("记加班 4h 昨天 设备调试", today=main.date(2026, 4, 7))
        self.assertEqual(parsed["action"], "overtime")
        self.assertEqual(str(parsed["date"]), "2026-04-06")
        self.assertEqual(parsed["hours"], 4.0)
        self.assertEqual(parsed["reason"], "设备调试")

    def test_parse_wechat_approval_command_supports_reason(self):
        parsed = main.parse_wechat_approval_command("不同意 工时不足")
        self.assertFalse(parsed["approved"])
        self.assertEqual(parsed["admin_comment"], "工时不足")

    def test_handle_wechat_text_command_creates_time_off_for_bound_user(self):
        member_id = self.create_member("member_a")
        with database.SessionLocal() as db:
            user = db.query(database.User).filter(database.User.id == member_id).first()
            user.wechat_user_id = "wx-member-a"
            db.commit()

            result = main.handle_wechat_text_command("wx-member-a", "调休 1h 2026-04-07 看牙", db)
            self.assertEqual(result["action"], "time_off")

            requests = db.query(database.TimeOffRequest).filter(database.TimeOffRequest.user_id == member_id).all()
            self.assertEqual(len(requests), 1)
            self.assertEqual(str(requests[0].date), "2026-04-07")
            self.assertEqual(requests[0].hours, 1.0)
            self.assertEqual(requests[0].reason, "看牙")
            self.assertEqual(requests[0].status, "pending")

    def test_handle_wechat_text_command_requires_bound_user(self):
        with database.SessionLocal() as db:
            with self.assertRaises(ValueError):
                main.handle_wechat_text_command("missing-user", "记加班 2h 今天 值班", db)

    def test_admin_wechat_approval_command_approves_latest_pending_request(self):
        member_id = self.create_member("member_a")
        with database.SessionLocal() as db:
            member = db.query(database.User).filter(database.User.id == member_id).first()
            admin = db.query(database.User).filter(database.User.name == "admin").first()
            admin.wechat_user_id = "wx-admin"
            member.wechat_user_id = "wx-member-a"

            request = database.TimeOffRequest(
                user_id=member_id,
                date=main.date(2026, 4, 7),
                hours=2.0,
                type="U",
                reason="看牙"
            )
            db.add(request)
            db.commit()

            result = main.handle_wechat_text_command("wx-admin", "同意 已处理", db)
            self.assertEqual(result["action"], "approval")
            self.assertEqual(result["target_type"], "time_off")

            db.refresh(request)
            self.assertEqual(request.status, "approved")
            self.assertEqual(request.admin_comment, "已处理")
            self.assertEqual(request.approved_by, admin.id)

    def test_monthly_export_generates_three_sheets_with_selected_order(self):
        member_a_id = self.create_member("member_a")
        member_b_id = self.create_member("member_b")

        response = self.client.post(
            "/api/export/monthly",
            json={"year": 2026, "month": 4, "user_ids": [member_b_id, member_a_id]},
            headers=self.admin_headers()
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response.headers["content-type"])

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["原始记录", "考勤统计", "数据统计"])
        self.assertEqual(workbook["原始记录"]["A2"].value, "member_b")
        self.assertEqual(workbook["原始记录"]["A3"].value, "member_a")
        self.assertEqual(workbook["考勤统计"]["A4"].value, 1)
        self.assertEqual(workbook["考勤统计"]["B4"].value, "member_b")
        self.assertEqual(workbook["考勤统计"]["A5"].value, 2)
        self.assertEqual(workbook["数据统计"]["B4"].value, "member_b")


if __name__ == "__main__":
    unittest.main()
